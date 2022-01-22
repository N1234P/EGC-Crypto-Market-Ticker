import requests
from openpyxl import load_workbook


def run_script(xlsx):
    wb = load_workbook(xlsx)

    sheet = wb["Sheet1"]

    ref = sheet["A1"]
    ref.value = "Wallet Addresses"

    ref = sheet["B1"]
    ref.value = "EGC Balances (SORTED)"

    holder_list = \
        requests.get("https://api.covalenthq.com/v1/56/tokens/0xC001BBe2B87079294C63EcE98BdD0a88D761434e/token_holders"
                     "/?quote-currency=USD&format=JSON&page-size=200000&key=ckey_76dca41983414ee1a5da0a1fb56").json()[
            'data']['items']

    rankings = [0, 0, 0, 0, 0, 0, 0, 0]

    for i in range(len(holder_list)):
        ref = sheet.cell(row=i + 2, column=1)
        ref.value = holder_list[i]['address']

        bal = int(holder_list[i]['balance'])
        ind = 11 - (20 - len(str(bal)))
        if len(str(bal)) > ind > 0:
            bal = int(str(bal)[:ind])

        ref = sheet.cell(row=i + 2, column=2)
        ref.value = bal

        if bal > 1000000000000:
            rankings[0] += 1
        elif bal > 500000000000:
            rankings[1] += 1
        elif bal > 100000000000:
            rankings[2] += 1
        elif bal > 50000000000:
            rankings[3] += 1
        elif bal > 10000000000:
            rankings[4] += 1
        elif bal > 1000000000:
            rankings[5] += 1
        elif bal > 100000000:
            rankings[6] += 1
        else:
            rankings[7] += 1
    # ----- largest
    ref = sheet.cell(row=len(holder_list), column=1)
    ref.value = "Holders with more than 1T EGC"
    ref = sheet.cell(row=len(holder_list), column=2)
    ref.value = rankings[0]
    ref = sheet.cell(row=len(holder_list), column=3)
    ref.value = str((rankings[0] / len(holder_list)) * 100) + " %"

    # ----- second largest
    ref = sheet.cell(row=len(holder_list) + 1, column=1)
    ref.value = "Holders between 500B-999B EGC"
    ref = sheet.cell(row=len(holder_list) + 1, column=2)
    ref.value = rankings[1]
    ref = sheet.cell(row=len(holder_list) + 1, column=3)
    ref.value = str((rankings[1] / len(holder_list)) * 100) + " %"

    # ----- third largest
    ref = sheet.cell(row=len(holder_list) + 2, column=1)
    ref.value = "Holders between 100B-499B EGC"
    ref = sheet.cell(row=len(holder_list) + 2, column=2)
    ref.value = rankings[2]
    ref = sheet.cell(row=len(holder_list) + 2, column=3)
    ref.value = str((rankings[2] / len(holder_list)) * 100) + " %"

    # ----- fourth largest
    ref = sheet.cell(row=len(holder_list) + 3, column=1)
    ref.value = "Holders between 50B-99B EGC"
    ref = sheet.cell(row=len(holder_list) + 3, column=2)
    ref.value = rankings[3]
    ref = sheet.cell(row=len(holder_list) + 3, column=3)
    ref.value = str((rankings[3] / len(holder_list)) * 100) + " %"

    # ----- fifth largest
    ref = sheet.cell(row=len(holder_list) + 4, column=1)
    ref.value = "Holders between 10B-49B EGC"
    ref = sheet.cell(row=len(holder_list) + 4, column=2)
    ref.value = rankings[4]
    ref = sheet.cell(row=len(holder_list) + 4, column=3)
    ref.value = str((rankings[4] / len(holder_list)) * 100) + " %"

    # ----- sixth largest
    ref = sheet.cell(row=len(holder_list) + 5, column=1)
    ref.value = "Holders between 1B-9B EGC"
    ref = sheet.cell(row=len(holder_list) + 5, column=2)
    ref.value = rankings[5]
    ref = sheet.cell(row=len(holder_list) + 5, column=3)
    ref.value = str((rankings[5] / len(holder_list)) * 100) + " %"

    # ----- seventh largest
    ref = sheet.cell(row=len(holder_list) + 6, column=1)
    ref.value = "Holders between 100M-999M EGC"
    ref = sheet.cell(row=len(holder_list) + 6, column=2)
    ref.value = rankings[6]
    ref = sheet.cell(row=len(holder_list) + 6, column=3)
    ref.value = str((rankings[6] / len(holder_list)) * 100) + " %"

    # ----- smallest
    ref = sheet.cell(row=len(holder_list) + 7, column=1)
    ref.value = "Holders with less than 100M EGC"
    ref = sheet.cell(row=len(holder_list) + 7, column=2)
    ref.value = rankings[7]
    ref = sheet.cell(row=len(holder_list) + 7, column=3)
    ref.value = str((rankings[7] / len(holder_list)) * 100) + " %"

    wb.save(xlsx)


run_script('EGC Holders.xlsx')